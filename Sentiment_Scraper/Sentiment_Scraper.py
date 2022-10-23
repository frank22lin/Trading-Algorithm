from datetime import date
import nest_asyncio; nest_asyncio.apply()  # This is needed to use sync API in repl
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

with sync_playwright() as pw:
    # create browser instance
    browser = pw.chromium.launch(
        # we can choose either a Headful (With GUI) or Headless mode:
        headless=True,
    )
    # create context
    # using context we can define page properties like viewport dimensions
    context = browser.new_context(
        # most common desktop viewport is 1920x1080
        viewport={"width": 1920, "height": 1080}
    )
    # create page aka browser tab which we'll be using to do everything
    page = context.new_page()
    page.goto("https://www.dailyfx.com/eur-usd")
    soup = BeautifulSoup(page.content(), "html.parser")
    results = soup.find(class_ = "dfx-rateDetail__percentageInfoText")

    percentage_long = results.attrs["data-value"]


wb = openpyxl.load_workbook(filename = 'scraped_sentiment_eurusd.xlsx')
sheet = wb.active
temp_max = sheet.max_row
sheet.cell(row = temp_max+1, column = 1).value = date.today()
sheet.cell(row = temp_max+1, column = 2).value = percentage_long
wb.save(filename = "scraped_sentiment_eurusd.xlsx")




# with pd.ExcelWriter('scraped_sentiment_eurusd.xlsx', engine='openpyxl', mode='a') as writer:
#     # df0 = pd.Series(date.today())
#     # df1 = pd.Series(percentage_long)

#     wb = writer.book
#     ws = wb.active

#     writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

#     writer.sheets[0].cell(row=ws.min_row, column=1).value = date.today()
#     writer.sheets[0].cell(row=ws.min_row, column=2).value = percentage_long

#     # df0.to_excel(writer, "Main", startrow=ws.min_row-1, startcol=0, index=False)
#     # df1.to_excel(writer, "Main", startrow=ws.min_row-1, startcol=1, index=False)
